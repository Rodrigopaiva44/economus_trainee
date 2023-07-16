from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# openpyxl is a famous libraly to work with excel
# openpyxl.utils is a group of tools to make the work easier

# Create a new Workbook
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Loading an existing workbook
wb = load_workbook('grades.xlsx')
ws = wb.active
ws = wb["Plan1"]  # to modify only this sheet

# Acessing cell values
print(ws["A2"].value)

# Change a value
ws["A2"].value = "Rodrigo"

# Save the workbook
wb.save("grades.xlsx")

# Acessing different sheets
print(wb.sheetnames)
print(wb["Plan1"])

# Create new sheets
wb.create_sheet("Test")
print(wb.sheetnames)

# Acessing multiple cells
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

# Merging cells
ws.merge_cells("A1:D2")
ws.unmerge_cells("A1:D2")

# Insert and delete rows
ws.insert_rows(7)
ws.delete_rows(7)
# Insert and delete colums
ws.insert_cols(2)
ws.delete_cols(2)

# copying and move cells
ws.move_range("C1:D11", rows=2, cols=2)
