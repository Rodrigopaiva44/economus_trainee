import openpyxl
import json


def create_excel_file(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Add data to the sheet
    for row_num, product in enumerate(data, 2):
        for col_num, value in enumerate(product.values(), 1):
            sheet.cell(row=row_num, column=col_num).value = value

    return workbook


def add_product(workbook, product):
    sheet = workbook.active
    new_row = [value for value in product.values()]
    sheet.append(new_row)


# JSON data
json_data = '''
  {
  "Produtos": [
    {
      "Nome": "Ração para cachorro/Kg",
      "Preço atual dos produtos": "R$ 19.90",
      "Preço ideal dos produtos": "R$ 17.00",
      "Custos Operacionais": "R$ 11.90",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Ração para gato/Kg",
      "Preço atual dos produtos": "R$ 20.70",
      "Preço ideal dos produtos": "R$ 19.86",
      "Custos Operacionais": "R$ 13.90",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Coleiras",
      "Preço atual dos produtos": "R$ 15.99",
      "Preço ideal dos produtos": "R$ 8.29",
      "Custos Operacionais": "R$ 5.80",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Brinquedos para animais",
      "Preço atual dos produtos": "R$ 26.30",
      "Preço ideal dos produtos": "R$ 21.86",
      "Custos Operacionais": "R$ 15.30",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Casinhas de animais",
      "Preço atual dos produtos": "R$ 30.90",
      "Preço ideal dos produtos": "R$ 26.86",
      "Custos Operacionais": "R$ 18.80",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Caminhas para animais",
      "Preço atual dos produtos": "R$ 35.20",
      "Preço ideal dos produtos": "R$ 31.29",
      "Custos Operacionais": "R$ 21.90",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Roupas para animais",
      "Preço atual dos produtos": "R$ 54.99",
      "Preço ideal dos produtos": "R$ 22.14",
      "Custos Operacionais": "R$ 15.50",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Produtos para higienização",
      "Preço atual dos produtos": "R$ 49.90",
      "Preço ideal dos produtos": "R$ 46.14",
      "Custos Operacionais": "R$ 32.30",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    },
    {
      "Nome": "Banho e Tosa",
      "Preço atual dos produtos": "R$ 70.00",
      "Preço ideal dos produtos": "R$ 81.86",
      "Custos Operacionais": "R$ 57.30",
      "Margem de Lucro": "30%",
      "Multiplicador": "70%"
    }
  ]
}
'''

# Convert JSON data to Python objects
data = json.loads(json_data)['Produtos']

# Create Excel file
workbook = create_excel_file(data)

# Save Excel file (./folder-name/file-name.xlsx)
workbook.save('Projetos/economus_trainee/help/product_data.xlsx')

# Example usage of add_product function
new_product = {
    "Nome": "Nova ração/Kg",
    "Preço atual dos produtos": "R$ 25.50",
    "Preço ideal dos produtos": "R$ 21.75",
    "Custos Operacionais": "R$ 15.80",
    "Margem de Lucro": "30%",
    "Multiplicador": "70%"
}

add_product(workbook, new_product)

# Save updated Excel file
workbook.save('Projetos/economus_trainee/help/product_data_updated.xlsx')
